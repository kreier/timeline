# Synchronize all dictionary CSVs into a single Excel workbook.
#
# Creates db/dictionary.xlsx with:
#   - a first worksheet "supported_languages"   -> content of supported_languages.csv
#   - a worksheet    "reference"                -> content of dictionary_reference.csv
#   - one worksheet per supported language      -> content of dictionary_XX.csv
#     (named by the two-letter language code XX, only for languages marked dict=TRUE)
#
# Formatting (frozen first row, column widths) is preserved from an existing
# dictionary.xlsx when present; new sheets are created when missing.

import os
import sys

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

DB_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_FILE = os.path.join(DB_DIR, "dictionary.xlsx")


def read_csv(name):
    path = os.path.join(DB_DIR, name)
    if not os.path.exists(path):
        print(f"WARNING: missing {name} - skipped")
        return None
    return pd.read_csv(path, dtype=str).fillna("")


def sheet_from_dataframe(ws, df):
    if df is None:
        return ws
    headers = list(df.columns)
    ws.append(headers)
    for _, row in df.iterrows():
        ws.append([row[col] for col in headers])
    return ws


def apply_formatting(ws, template_ws):
    # Freeze the first row so the header stays visible when scrolling.
    if template_ws is not None and template_ws.freeze_panes:
        ws.freeze_panes = template_ws.freeze_panes
    else:
        ws.freeze_panes = "A2"

    # Reuse column widths from the template sheet; give new columns a default width.
    widths = {}
    if template_ws is not None:
        for col_letter, dim in template_ws.column_dimensions.items():
            if dim.width:
                widths[col_letter] = dim.width
    for col_idx in range(1, (ws.max_column or 1) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = widths.get(col_letter, 20)


def supported_languages():
    df = read_csv("supported_languages.csv")
    if df is None:
        return []
    if "dict" not in df.columns or "key" not in df.columns:
        return []
    true_values = {"TRUE", "True", "true", "1", "1.0"}
    codes = [str(k).strip() for k, v in zip(df["key"], df["dict"]) if str(v).strip() in true_values]
    return [c for c in codes if c and c.lower() != "nan"]


def main():
    # Load the existing workbook as a formatting template if it exists.
    template_ws = None
    if os.path.exists(OUTPUT_FILE):
        try:
            template_book = load_workbook(OUTPUT_FILE)
            template_ws = template_book[template_book.sheetnames[0]]
        except Exception as exc:
            print(f"WARNING: could not reuse formatting from {OUTPUT_FILE}: {exc}")

    book = Workbook()
    book.remove(book.active)  # drop the default empty sheet

    # 1) supported_languages -> first worksheet
    ws = book.create_sheet("supported_languages")
    sheet_from_dataframe(ws, read_csv("supported_languages.csv"))
    apply_formatting(ws, template_ws)

    # 2) dictionary_reference -> reference worksheet
    ws = book.create_sheet("reference")
    sheet_from_dataframe(ws, read_csv("dictionary_reference.csv"))
    apply_formatting(ws, template_ws)

    # 3) one worksheet per supported language dictionary
    for code in supported_languages():
        df = read_csv(f"dictionary_{code}.csv")
        if df is None:
            print(f"NOTE: no dictionary_{code}.csv for supported language '{code}' - skipped")
            continue
        ws = book.create_sheet(code)
        sheet_from_dataframe(ws, df)
        apply_formatting(ws, template_ws)

    book.save(OUTPUT_FILE)
    print(f"Wrote {OUTPUT_FILE} with sheets: {book.sheetnames}")


if __name__ == "__main__":
    sys.exit(main())
